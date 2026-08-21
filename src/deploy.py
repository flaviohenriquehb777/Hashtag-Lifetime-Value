"""
deploy.py
=========
Geração e validação de uma calculadora de LTV em Excel baseada no
modelo final de Regressão Linear.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.model_selection import train_test_split
try:
    from win32com.client import Dispatch
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

from .models import build_gemini_style_pipelines, evaluate_models_on_final_base_test


MODEL_NAME = "2. Regressão Linear"
RAW_INPUT_COLUMNS = [
    "valor_1_compra",
    "recorrente_1_compra",
    "Produto Fonte",
    "Fonte Campanha",
    "Sexo",
    "Formacao",
    "mes_compra",
    "dia_semana_compra",
]
INPUT_DESCRIPTIONS = {
    "valor_1_compra": "Valor da primeira compra do cliente, em R$.",
    "recorrente_1_compra": "Indicador binário da base tratada: 0 = não recorrente, 1 = recorrente.",
    "Produto Fonte": "Produto de entrada do cliente na jornada comercial.",
    "Fonte Campanha": "Origem/canal de aquisição atribuído ao cliente.",
    "Sexo": "Categoria de sexo consolidada na etapa de limpeza.",
    "Formacao": "Escolaridade consolidada na etapa de limpeza.",
    "mes_compra": "Mês da compra extraído da data original (1 a 12).",
    "dia_semana_compra": "Dia da semana da compra extraído da data original (0 = segunda, 6 = domingo).",
}
BASELINES = {
    "Produto Fonte": "Análise de Dados",
    "Fonte Campanha": "Checkout",
    "Sexo": "Feminino",
    "Formacao": "Fundamental",
    "dia_semana_compra": 0,
    "mes_compra": 1,
}
INPUT_NAME_MAP = {
    "valor_1_compra": "Valor1Compra_Input",
    "recorrente_1_compra": "Recorrente1Compra_Input",
    "Produto Fonte": "ProdutoFonte_Input",
    "Fonte Campanha": "FonteCampanha_Input",
    "Sexo": "Sexo_Input",
    "Formacao": "Formacao_Input",
    "mes_compra": "MesCompra_Input",
    "dia_semana_compra": "DiaSemanaCompra_Input",
    "cac_real": "CACReal_Input",
    "ltv_cac_ratio": "Razao_LTV_CAC_Alvo",
    "ltv_predito": "LTV_Previsto",
    "cac_maximo": "CAC_Maximo_Recomendado",
    "status": "Status_Aquisicao",
}


@dataclass
class LinearCalculatorArtifacts:
    intercept: float
    feature_names: list[str]
    coefficients: list[float]
    scaler_mean: float
    scaler_scale: float
    coef_valor_1_compra: float
    coef_recorrente_1_compra: float
    coef_produto_fonte: dict[object, float]
    coef_fonte_campanha: dict[object, float]
    coef_sexo: dict[object, float]
    coef_formacao: dict[object, float]
    coef_mes_compra: dict[object, float]
    coef_dia_semana_compra: dict[object, float]
    allowed_categories: dict[str, list[object]]
    training_ranges: dict[str, tuple[object, object]]
    metrics: dict[str, float]
    split_sizes: dict[str, int]
    training_rows: int
    test_rows: int


def _to_python_scalar(value: object) -> object:
    if hasattr(value, "item"):
        return value.item()
    return value


def _extract_categorical_maps(
    categories: list[list[object]],
    coefficients: list[float],
) -> tuple[dict[str, dict[object, float]], int]:
    column_names = [
        "Produto Fonte",
        "Fonte Campanha",
        "Sexo",
        "Formacao",
        "dia_semana_compra",
        "mes_compra",
    ]
    maps: dict[str, dict[object, float]] = {}
    start = 1  # índice 0 é valor_1_compra padronizado

    for column_name, column_categories in zip(column_names, categories, strict=True):
        encoded_categories = [_to_python_scalar(value) for value in column_categories[1:]]
        end = start + len(encoded_categories)
        maps[column_name] = dict(zip(encoded_categories, coefficients[start:end], strict=True))
        start = end

    return maps, start


def _normalize_mapping_keys(mapping: dict[object, float], *, as_int: bool = False) -> dict[object, float]:
    normalized: dict[object, float] = {}
    for key, value in mapping.items():
        normalized_key = int(key) if as_int else key
        normalized[normalized_key] = value
    return normalized


def fit_linear_calculator_artifacts(
    df: pd.DataFrame,
    *,
    target_col: str = "LTV",
    test_size: float = 0.2,
    random_state: int = 42,
) -> LinearCalculatorArtifacts:
    """Treina o pipeline linear final e extrai todos os artefatos para o Excel."""
    X = df.drop(columns=[target_col]).copy()
    y = df[target_col].copy()

    X_train, X_test, y_train, y_test = train_test_split(
        X,
        y,
        test_size=test_size,
        random_state=random_state,
    )

    pipeline = build_gemini_style_pipelines()[MODEL_NAME]
    pipeline.fit(X_train, y_train)

    preprocessor = pipeline.named_steps["prep"]
    model = pipeline.named_steps["model"]
    ohe = preprocessor.named_transformers_["cat"]
    scaler = preprocessor.named_transformers_["num"]

    feature_names = preprocessor.get_feature_names_out().tolist()
    coefficients = model.coef_.tolist()
    intercept = float(model.intercept_)

    categorical_categories = [
        [_to_python_scalar(value) for value in category_list]
        for category_list in ohe.categories_
    ]
    coefficient_maps, end_index = _extract_categorical_maps(categorical_categories, coefficients)
    coef_recorrente = float(coefficients[end_index])

    metrics_df, split_info = evaluate_models_on_final_base_test(
        df,
        target_col=target_col,
        test_size=test_size,
        random_state=random_state,
    )
    linear_metrics = metrics_df.loc[metrics_df["modelo"] == MODEL_NAME].iloc[0]

    allowed_categories = {
        "Produto Fonte": sorted(X_train["Produto Fonte"].dropna().unique().tolist()),
        "Fonte Campanha": sorted(X_train["Fonte Campanha"].dropna().unique().tolist()),
        "Sexo": sorted(X_train["Sexo"].dropna().unique().tolist()),
        "Formacao": sorted(X_train["Formacao"].dropna().unique().tolist()),
        "mes_compra": sorted(int(value) for value in X_train["mes_compra"].dropna().unique().tolist()),
        "dia_semana_compra": sorted(int(value) for value in X_train["dia_semana_compra"].dropna().unique().tolist()),
        "recorrente_1_compra": [0, 1],
    }
    training_ranges = {
        "valor_1_compra": (
            float(X_train["valor_1_compra"].min()),
            float(X_train["valor_1_compra"].max()),
        ),
        "mes_compra": (
            int(min(allowed_categories["mes_compra"])),
            int(max(allowed_categories["mes_compra"])),
        ),
        "dia_semana_compra": (
            int(min(allowed_categories["dia_semana_compra"])),
            int(max(allowed_categories["dia_semana_compra"])),
        ),
    }

    coef_mes_compra = _normalize_mapping_keys(coefficient_maps["mes_compra"], as_int=True)
    coef_dia_semana_compra = _normalize_mapping_keys(
        coefficient_maps["dia_semana_compra"],
        as_int=True,
    )

    return LinearCalculatorArtifacts(
        intercept=intercept,
        feature_names=feature_names,
        coefficients=[float(value) for value in coefficients],
        scaler_mean=float(scaler.mean_[0]),
        scaler_scale=float(scaler.scale_[0]),
        coef_valor_1_compra=float(coefficients[0]),
        coef_recorrente_1_compra=coef_recorrente,
        coef_produto_fonte=coefficient_maps["Produto Fonte"],
        coef_fonte_campanha=coefficient_maps["Fonte Campanha"],
        coef_sexo=coefficient_maps["Sexo"],
        coef_formacao=coefficient_maps["Formacao"],
        coef_mes_compra=coef_mes_compra,
        coef_dia_semana_compra=coef_dia_semana_compra,
        allowed_categories=allowed_categories,
        training_ranges=training_ranges,
        metrics={
            "R2": float(linear_metrics["R2"]),
            "RMSE": float(linear_metrics["RMSE"]),
            "MAE": float(linear_metrics["MAE"]),
        },
        split_sizes={
            row["split"]: int(row["linhas"])
            for row in split_info.to_dict(orient="records")
        },
        training_rows=int(len(X_train)),
        test_rows=int(len(X_test)),
    )


def predict_linear_formula_python(
    artifacts: LinearCalculatorArtifacts,
    client_inputs: dict[str, object],
) -> float:
    """Aplica em Python exatamente a mesma fórmula que será escrita no Excel."""
    valor_padronizado = (
        float(client_inputs["valor_1_compra"]) - artifacts.scaler_mean
    ) / artifacts.scaler_scale

    prediction = artifacts.intercept
    prediction += artifacts.coef_valor_1_compra * valor_padronizado
    prediction += artifacts.coef_recorrente_1_compra * int(client_inputs["recorrente_1_compra"])
    prediction += artifacts.coef_produto_fonte.get(client_inputs["Produto Fonte"], 0.0)
    prediction += artifacts.coef_fonte_campanha.get(client_inputs["Fonte Campanha"], 0.0)
    prediction += artifacts.coef_sexo.get(client_inputs["Sexo"], 0.0)
    prediction += artifacts.coef_formacao.get(client_inputs["Formacao"], 0.0)
    prediction += artifacts.coef_mes_compra.get(int(client_inputs["mes_compra"]), 0.0)
    prediction += artifacts.coef_dia_semana_compra.get(int(client_inputs["dia_semana_compra"]), 0.0)
    return float(prediction)


def default_test_clients() -> list[dict[str, object]]:
    """Retorna 3 clientes fictícios para validar Python vs Excel."""
    return [
        {
            "cliente_teste": "Cliente A",
            "valor_1_compra": 997.00,
            "recorrente_1_compra": 0,
            "Produto Fonte": "Power BI",
            "Fonte Campanha": "Lista Espera",
            "Sexo": "Outros",
            "Formacao": "Não informado",
            "mes_compra": 1,
            "dia_semana_compra": 6,
            "cac_real": 500.00,
            "ltv_cac_ratio": 3.0,
        },
        {
            "cliente_teste": "Cliente B",
            "valor_1_compra": 297.00,
            "recorrente_1_compra": 1,
            "Produto Fonte": "Excel",
            "Fonte Campanha": "Webinar",
            "Sexo": "Feminino",
            "Formacao": "Médio",
            "mes_compra": 5,
            "dia_semana_compra": 2,
            "cac_real": 700.00,
            "ltv_cac_ratio": 3.0,
        },
        {
            "cliente_teste": "Cliente C",
            "valor_1_compra": 1_499.00,
            "recorrente_1_compra": 1,
            "Produto Fonte": "Vitalício",
            "Fonte Campanha": "Comercial",
            "Sexo": "Masculino",
            "Formacao": "Superior+",
            "mes_compra": 11,
            "dia_semana_compra": 4,
            "cac_real": 1_200.00,
            "ltv_cac_ratio": 3.0,
        },
    ]


def _add_named_range(workbook: Workbook, name: str, target: str) -> None:
    workbook.defined_names.add(DefinedName(name=name, attr_text=target))


def _write_vertical_block(
    ws,
    start_row: int,
    start_col: int,
    values: list[object],
) -> tuple[int, int]:
    for offset, value in enumerate(values):
        ws.cell(row=start_row + offset, column=start_col, value=value)
    return start_row, start_row + len(values) - 1


def create_ltv_calculator_workbook(
    artifacts: LinearCalculatorArtifacts,
    output_path: str | Path,
) -> Path:
    """Gera o arquivo .xlsx completo da calculadora de LTV."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws_calc = wb.active
    ws_calc.title = "Calculadora"
    ws_aux = wb.create_sheet("Apoio")
    ws_doc = wb.create_sheet("Documentacao")

    # Estilo base
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    # Aba Calculadora
    ws_calc["A1"] = "Calculadora de LTV"
    ws_calc["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws_calc["A1"].fill = title_fill
    ws_calc.merge_cells("A1:D1")

    labels = [
        ("A4", "Valor da 1ª compra (R$)", "B4", 500.00),
        ("A5", "Recorrente na 1ª compra (0/1)", "B5", 0),
        ("A6", "Produto Fonte", "B6", artifacts.allowed_categories["Produto Fonte"][0]),
        ("A7", "Fonte Campanha", "B7", artifacts.allowed_categories["Fonte Campanha"][0]),
        ("A8", "Sexo", "B8", artifacts.allowed_categories["Sexo"][0]),
        ("A9", "Formação", "B9", artifacts.allowed_categories["Formacao"][0]),
        ("A10", "Mês da compra", "B10", artifacts.allowed_categories["mes_compra"][0]),
        ("A11", "Dia da semana da compra", "B11", artifacts.allowed_categories["dia_semana_compra"][0]),
        ("A12", "CAC Real (R$)", "B12", 0.0),
        ("A13", "Razão LTV:CAC alvo", "B13", 3.0),
    ]

    for label_cell, label, value_cell, default_value in labels:
        ws_calc[label_cell] = label
        ws_calc[label_cell].fill = section_fill
        ws_calc[label_cell].font = Font(bold=True)
        ws_calc[label_cell].border = border
        ws_calc[value_cell] = default_value
        ws_calc[value_cell].fill = input_fill
        ws_calc[value_cell].border = border
        ws_calc[value_cell].protection = Protection(locked=False)

    ws_calc["A15"] = "LTV Previsto (R$)"
    ws_calc["A16"] = "CAC Máximo Recomendado (R$)"
    ws_calc["A17"] = "Status da aquisição"
    for cell in ("A15", "A16", "A17"):
        ws_calc[cell].fill = section_fill
        ws_calc[cell].font = Font(bold=True)
        ws_calc[cell].border = border

    # Apoio: listas, coeficientes e metadados
    ws_aux["A1"] = "Listas para validação"
    ws_aux["J1"] = "Coeficientes do modelo"
    ws_aux["AB1"] = "Metadados"

    row_start = 2
    product_rows = _write_vertical_block(ws_aux, row_start, 1, artifacts.allowed_categories["Produto Fonte"])
    source_rows = _write_vertical_block(ws_aux, row_start, 2, artifacts.allowed_categories["Fonte Campanha"])
    sex_rows = _write_vertical_block(ws_aux, row_start, 3, artifacts.allowed_categories["Sexo"])
    education_rows = _write_vertical_block(ws_aux, row_start, 4, artifacts.allowed_categories["Formacao"])
    month_rows = _write_vertical_block(ws_aux, row_start, 5, artifacts.allowed_categories["mes_compra"])
    day_rows = _write_vertical_block(ws_aux, row_start, 6, artifacts.allowed_categories["dia_semana_compra"])

    # Categorias codificadas e coeficientes
    ws_aux["J2"] = "Produto Fonte"
    ws_aux["K2"] = "Coef"
    ws_aux["M2"] = "Fonte Campanha"
    ws_aux["N2"] = "Coef"
    ws_aux["P2"] = "Sexo"
    ws_aux["Q2"] = "Coef"
    ws_aux["S2"] = "Formação"
    ws_aux["T2"] = "Coef"
    ws_aux["V2"] = "Mês"
    ws_aux["W2"] = "Coef"
    ws_aux["Y2"] = "Dia semana"
    ws_aux["Z2"] = "Coef"

    def _write_map(start_row: int, category_col: int, coef_col: int, mapping: dict[object, float]) -> tuple[int, int]:
        for idx, (key, value) in enumerate(mapping.items(), start=start_row):
            ws_aux.cell(row=idx, column=category_col, value=key)
            ws_aux.cell(row=idx, column=coef_col, value=float(value))
        return start_row, start_row + len(mapping) - 1

    product_coef_rows = _write_map(3, 10, 11, artifacts.coef_produto_fonte)
    source_coef_rows = _write_map(3, 13, 14, artifacts.coef_fonte_campanha)
    sex_coef_rows = _write_map(3, 16, 17, artifacts.coef_sexo)
    education_coef_rows = _write_map(3, 19, 20, artifacts.coef_formacao)
    month_coef_rows = _write_map(3, 22, 23, artifacts.coef_mes_compra)
    day_coef_rows = _write_map(3, 25, 26, artifacts.coef_dia_semana_compra)

    metadata_pairs = [
        ("Intercepto", artifacts.intercept),
        ("Coef_valor_1_compra", artifacts.coef_valor_1_compra),
        ("Media_valor_1_compra", artifacts.scaler_mean),
        ("Escala_valor_1_compra", artifacts.scaler_scale),
        ("Coef_recorrente_1_compra", artifacts.coef_recorrente_1_compra),
        ("R2", artifacts.metrics["R2"]),
        ("RMSE", artifacts.metrics["RMSE"]),
        ("MAE", artifacts.metrics["MAE"]),
    ]
    for idx, (label, value) in enumerate(metadata_pairs, start=2):
        ws_aux.cell(row=idx, column=28, value=label)
        ws_aux.cell(row=idx, column=29, value=value)

    # Named ranges: inputs e outputs
    for raw_name, excel_name in INPUT_NAME_MAP.items():
        if raw_name in {"cac_real", "ltv_cac_ratio", "ltv_predito", "cac_maximo", "status"}:
            continue
    _add_named_range(wb, "Valor1Compra_Input", "'Calculadora'!$B$4")
    _add_named_range(wb, "Recorrente1Compra_Input", "'Calculadora'!$B$5")
    _add_named_range(wb, "ProdutoFonte_Input", "'Calculadora'!$B$6")
    _add_named_range(wb, "FonteCampanha_Input", "'Calculadora'!$B$7")
    _add_named_range(wb, "Sexo_Input", "'Calculadora'!$B$8")
    _add_named_range(wb, "Formacao_Input", "'Calculadora'!$B$9")
    _add_named_range(wb, "MesCompra_Input", "'Calculadora'!$B$10")
    _add_named_range(wb, "DiaSemanaCompra_Input", "'Calculadora'!$B$11")
    _add_named_range(wb, "CACReal_Input", "'Calculadora'!$B$12")
    _add_named_range(wb, "Razao_LTV_CAC_Alvo", "'Calculadora'!$B$13")
    _add_named_range(wb, "LTV_Previsto", "'Calculadora'!$B$15")
    _add_named_range(wb, "CAC_Maximo_Recomendado", "'Calculadora'!$B$16")
    _add_named_range(wb, "Status_Aquisicao", "'Calculadora'!$B$17")

    # Named ranges: apoio
    _add_named_range(wb, "Lista_Produto_Fonte", f"'Apoio'!$A${product_rows[0]}:$A${product_rows[1]}")
    _add_named_range(wb, "Lista_Fonte_Campanha", f"'Apoio'!$B${source_rows[0]}:$B${source_rows[1]}")
    _add_named_range(wb, "Lista_Sexo", f"'Apoio'!$C${sex_rows[0]}:$C${sex_rows[1]}")
    _add_named_range(wb, "Lista_Formacao", f"'Apoio'!$D${education_rows[0]}:$D${education_rows[1]}")
    _add_named_range(wb, "Lista_Mes_Compra", f"'Apoio'!$E${month_rows[0]}:$E${month_rows[1]}")
    _add_named_range(wb, "Lista_Dia_Semana", f"'Apoio'!$F${day_rows[0]}:$F${day_rows[1]}")

    _add_named_range(wb, "Categorias_Produto_Fonte_Modelo", f"'Apoio'!$J${product_coef_rows[0]}:$J${product_coef_rows[1]}")
    _add_named_range(wb, "Coef_Produto_Fonte", f"'Apoio'!$K${product_coef_rows[0]}:$K${product_coef_rows[1]}")
    _add_named_range(wb, "Categorias_Fonte_Campanha_Modelo", f"'Apoio'!$M${source_coef_rows[0]}:$M${source_coef_rows[1]}")
    _add_named_range(wb, "Coef_Fonte_Campanha", f"'Apoio'!$N${source_coef_rows[0]}:$N${source_coef_rows[1]}")
    _add_named_range(wb, "Categorias_Sexo_Modelo", f"'Apoio'!$P${sex_coef_rows[0]}:$P${sex_coef_rows[1]}")
    _add_named_range(wb, "Coef_Sexo", f"'Apoio'!$Q${sex_coef_rows[0]}:$Q${sex_coef_rows[1]}")
    _add_named_range(wb, "Categorias_Formacao_Modelo", f"'Apoio'!$S${education_coef_rows[0]}:$S${education_coef_rows[1]}")
    _add_named_range(wb, "Coef_Formacao", f"'Apoio'!$T${education_coef_rows[0]}:$T${education_coef_rows[1]}")
    _add_named_range(wb, "Categorias_Mes_Modelo", f"'Apoio'!$V${month_coef_rows[0]}:$V${month_coef_rows[1]}")
    _add_named_range(wb, "Coef_Mes_Compra", f"'Apoio'!$W${month_coef_rows[0]}:$W${month_coef_rows[1]}")
    _add_named_range(wb, "Categorias_Dia_Semana_Modelo", f"'Apoio'!$Y${day_coef_rows[0]}:$Y${day_coef_rows[1]}")
    _add_named_range(wb, "Coef_Dia_Semana", f"'Apoio'!$Z${day_coef_rows[0]}:$Z${day_coef_rows[1]}")

    _add_named_range(wb, "Intercepto_Modelo", "'Apoio'!$AC$2")
    _add_named_range(wb, "Coef_Valor1Compra", "'Apoio'!$AC$3")
    _add_named_range(wb, "Valor1Compra_Mean", "'Apoio'!$AC$4")
    _add_named_range(wb, "Valor1Compra_Scale", "'Apoio'!$AC$5")
    _add_named_range(wb, "Coef_Recorrente1Compra", "'Apoio'!$AC$6")

    # Helper visível só via named range
    ws_calc["J2"] = "Valor1Compra_Z"
    ws_calc["K2"] = "=(Valor1Compra_Input-Valor1Compra_Mean)/Valor1Compra_Scale"
    _add_named_range(wb, "Valor1Compra_Z", "'Calculadora'!$K$2")
    ws_calc.column_dimensions["J"].hidden = True
    ws_calc.column_dimensions["K"].hidden = True

    formula_ltv = (
        "=Intercepto_Modelo"
        "+Coef_Valor1Compra*Valor1Compra_Z"
        "+Coef_Recorrente1Compra*Recorrente1Compra_Input"
        "+SUMPRODUCT(--(ProdutoFonte_Input=Categorias_Produto_Fonte_Modelo),Coef_Produto_Fonte)"
        "+SUMPRODUCT(--(FonteCampanha_Input=Categorias_Fonte_Campanha_Modelo),Coef_Fonte_Campanha)"
        "+SUMPRODUCT(--(Sexo_Input=Categorias_Sexo_Modelo),Coef_Sexo)"
        "+SUMPRODUCT(--(Formacao_Input=Categorias_Formacao_Modelo),Coef_Formacao)"
        "+SUMPRODUCT(--(MesCompra_Input=Categorias_Mes_Modelo),Coef_Mes_Compra)"
        "+SUMPRODUCT(--(DiaSemanaCompra_Input=Categorias_Dia_Semana_Modelo),Coef_Dia_Semana)"
    )
    ws_calc["B15"] = formula_ltv
    ws_calc["B16"] = "=IFERROR(LTV_Previsto/Razao_LTV_CAC_Alvo,\"\")"
    ws_calc["B17"] = (
        '=IF(CACReal_Input="",'
        '"Sem CAC informado",'
        'IF(CACReal_Input<=CAC_Maximo_Recomendado*0.9,"Saudável",'
        'IF(CACReal_Input<=CAC_Maximo_Recomendado,"No limite","Acima do recomendado")))'
    )
    for cell in ("B15", "B16", "B17"):
        ws_calc[cell].border = border
        ws_calc[cell].font = Font(bold=True)

    ws_calc["B15"].number_format = 'R$ #,##0.00'
    ws_calc["B16"].number_format = 'R$ #,##0.00'
    ws_calc["B12"].number_format = 'R$ #,##0.00'
    ws_calc["B4"].number_format = 'R$ #,##0.00'

    # Validação de dados
    validations = [
        (DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False), "B4"),
        (DataValidation(type="list", formula1='"0,1"', allow_blank=False), "B5"),
        (DataValidation(type="list", formula1="=Lista_Produto_Fonte", allow_blank=False), "B6"),
        (DataValidation(type="list", formula1="=Lista_Fonte_Campanha", allow_blank=False), "B7"),
        (DataValidation(type="list", formula1="=Lista_Sexo", allow_blank=False), "B8"),
        (DataValidation(type="list", formula1="=Lista_Formacao", allow_blank=False), "B9"),
        (DataValidation(type="list", formula1="=Lista_Mes_Compra", allow_blank=False), "B10"),
        (DataValidation(type="list", formula1="=Lista_Dia_Semana", allow_blank=False), "B11"),
        (DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True), "B12"),
        (DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False), "B13"),
    ]
    for validation, cell in validations:
        ws_calc.add_data_validation(validation)
        validation.add(ws_calc[cell])

    # Formatação condicional
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    ws_calc.conditional_formatting.add(
        "B17",
        FormulaRule(formula=['$B$17="Saudável"'], fill=green_fill),
    )
    ws_calc.conditional_formatting.add(
        "B17",
        FormulaRule(formula=['$B$17="No limite"'], fill=yellow_fill),
    )
    ws_calc.conditional_formatting.add(
        "B17",
        FormulaRule(formula=['$B$17="Acima do recomendado"'], fill=red_fill),
    )

    # Proteção
    ws_calc.protection.sheet = True
    ws_calc.protection.password = "ltv"
    ws_aux.protection.sheet = True
    ws_aux.protection.password = "ltv"
    ws_aux.sheet_state = "hidden"

    # Aba Documentacao
    ws_doc["A1"] = "Documentação da calculadora"
    ws_doc["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_doc["A1"].fill = title_fill
    ws_doc.merge_cells("A1:F1")

    doc_lines = [
        ("Modelo escolhido", "Regressão Linear"),
        ("R2", artifacts.metrics["R2"]),
        ("RMSE", artifacts.metrics["RMSE"]),
        ("MAE", artifacts.metrics["MAE"]),
        ("Intercepto", artifacts.intercept),
        ("Linhas de treino", artifacts.training_rows),
        ("Linhas de teste", artifacts.test_rows),
        ("Período original da base", "01/01/2023 a 31/05/2024"),
        ("Observação", "Não usar fora das faixas e categorias observadas sem revalidação do modelo."),
    ]
    for idx, (label, value) in enumerate(doc_lines, start=3):
        ws_doc[f"A{idx}"] = label
        ws_doc[f"B{idx}"] = value

    ws_doc["A14"] = "Variáveis de input"
    ws_doc["A14"].font = Font(bold=True)
    ws_doc["A15"] = "Variável"
    ws_doc["B15"] = "Descrição"
    ws_doc["C15"] = "Faixa / categorias observadas no treino"
    ws_doc["D15"] = "Baseline do modelo"
    for cell in ("A15", "B15", "C15", "D15"):
        ws_doc[cell].font = Font(bold=True)
        ws_doc[cell].fill = section_fill

    raw_range_text = {
        "valor_1_compra": f"{artifacts.training_ranges['valor_1_compra'][0]:,.2f} a {artifacts.training_ranges['valor_1_compra'][1]:,.2f}",
        "recorrente_1_compra": "0 ou 1",
        "Produto Fonte": ", ".join(map(str, artifacts.allowed_categories["Produto Fonte"])),
        "Fonte Campanha": ", ".join(map(str, artifacts.allowed_categories["Fonte Campanha"])),
        "Sexo": ", ".join(map(str, artifacts.allowed_categories["Sexo"])),
        "Formacao": ", ".join(map(str, artifacts.allowed_categories["Formacao"])),
        "mes_compra": f"{artifacts.training_ranges['mes_compra'][0]} a {artifacts.training_ranges['mes_compra'][1]}",
        "dia_semana_compra": f"{artifacts.training_ranges['dia_semana_compra'][0]} a {artifacts.training_ranges['dia_semana_compra'][1]} (0=segunda, 6=domingo)",
    }
    for idx, column_name in enumerate(RAW_INPUT_COLUMNS, start=16):
        ws_doc[f"A{idx}"] = column_name
        ws_doc[f"B{idx}"] = INPUT_DESCRIPTIONS[column_name]
        ws_doc[f"C{idx}"] = raw_range_text[column_name]
        ws_doc[f"D{idx}"] = BASELINES.get(column_name, "-")

    ws_doc["A27"] = "Features transformadas e coeficientes (ordem exata)"
    ws_doc["A27"].font = Font(bold=True)
    ws_doc["A28"] = "Feature"
    ws_doc["B28"] = "Coeficiente"
    ws_doc["A28"].font = Font(bold=True)
    ws_doc["B28"].font = Font(bold=True)
    ws_doc["A28"].fill = section_fill
    ws_doc["B28"].fill = section_fill
    for idx, (feature_name, coef) in enumerate(zip(artifacts.feature_names, artifacts.coefficients, strict=True), start=29):
        ws_doc[f"A{idx}"] = feature_name
        ws_doc[f"B{idx}"] = coef

    # Layout
    for ws in (ws_calc, ws_doc):
        for column in range(1, 7):
            ws.column_dimensions[get_column_letter(column)].width = 24
    ws_calc.column_dimensions["A"].width = 32
    ws_calc.column_dimensions["B"].width = 24
    ws_doc.column_dimensions["A"].width = 28
    ws_doc.column_dimensions["B"].width = 45
    ws_doc.column_dimensions["C"].width = 45
    ws_doc.column_dimensions["D"].width = 24
    ws_doc.freeze_panes = "A15"
    for row in ws_doc.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output)
    return output


def _set_named_range_value(workbook, name: str, value: object) -> None:
    workbook.Names(name).RefersToRange.Value = value


def _get_named_range_value(workbook, name: str) -> object:
    return workbook.Names(name).RefersToRange.Value


def validate_excel_against_python(
    workbook_path: str | Path,
    artifacts: LinearCalculatorArtifacts,
    test_clients: list[dict[str, object]],
    *,
    tolerance: float = 1e-9,
) -> pd.DataFrame:
    """Executa o loop de validação comparando Excel real vs fórmula Python."""
    if not HAS_WIN32COM:
        raise ImportError(
            "O módulo 'pywin32' é necessário para a validação real no Excel. "
            "Esta funcionalidade só está disponível em Windows com Excel instalado."
        )

    workbook_path = str(Path(workbook_path).resolve())
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    workbook = excel.Workbooks.Open(workbook_path)
    results: list[dict[str, object]] = []

    try:
        for client in test_clients:
            _set_named_range_value(workbook, "Valor1Compra_Input", float(client["valor_1_compra"]))
            _set_named_range_value(workbook, "Recorrente1Compra_Input", int(client["recorrente_1_compra"]))
            _set_named_range_value(workbook, "ProdutoFonte_Input", str(client["Produto Fonte"]))
            _set_named_range_value(workbook, "FonteCampanha_Input", str(client["Fonte Campanha"]))
            _set_named_range_value(workbook, "Sexo_Input", str(client["Sexo"]))
            _set_named_range_value(workbook, "Formacao_Input", str(client["Formacao"]))
            _set_named_range_value(workbook, "MesCompra_Input", int(client["mes_compra"]))
            _set_named_range_value(workbook, "DiaSemanaCompra_Input", int(client["dia_semana_compra"]))
            _set_named_range_value(workbook, "CACReal_Input", float(client["cac_real"]))
            _set_named_range_value(workbook, "Razao_LTV_CAC_Alvo", float(client["ltv_cac_ratio"]))

            excel.CalculateFullRebuild()

            ltv_python = predict_linear_formula_python(artifacts, client)
            ltv_excel = float(_get_named_range_value(workbook, "LTV_Previsto"))
            cac_max_excel = float(_get_named_range_value(workbook, "CAC_Maximo_Recomendado"))

            if abs(ltv_python - ltv_excel) > tolerance:
                raise ValueError(
                    "Diferença encontrada entre Python e Excel "
                    f"para {client['cliente_teste']}: Python={ltv_python}, Excel={ltv_excel}"
                )

            input_summary = (
                f"valor_1_compra={client['valor_1_compra']}; "
                f"recorrente_1_compra={client['recorrente_1_compra']}; "
                f"Produto Fonte={client['Produto Fonte']}; "
                f"Fonte Campanha={client['Fonte Campanha']}; "
                f"Sexo={client['Sexo']}; "
                f"Formacao={client['Formacao']}; "
                f"mes_compra={client['mes_compra']}; "
                f"dia_semana_compra={client['dia_semana_compra']}"
            )
            results.append(
                {
                    "cliente_teste": client["cliente_teste"],
                    "inputs": input_summary,
                    "LTV (Python)": ltv_python,
                    "LTV (Excel)": ltv_excel,
                    "CAC máximo recomendado": cac_max_excel,
                }
            )

        workbook.Save()
    finally:
        workbook.Close(SaveChanges=True)
        excel.Quit()

    return pd.DataFrame(results)

