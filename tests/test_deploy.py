from __future__ import annotations

import pytest
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook

from src.deploy import (
    create_ltv_calculator_workbook,
    default_test_clients,
    fit_linear_calculator_artifacts,
    predict_linear_formula_python,
)

@pytest.fixture
def dummy_final_df():
    """Gera um DataFrame final para testes de deploy."""
    n = 500
    dates = pd.date_range("2023-01-01", periods=n, freq="D")
    df = pd.DataFrame({
        "LTV": np.random.uniform(100, 1000, n),
        "data_compra": dates,
        "valor_1_compra": np.random.uniform(50, 500, n),
        "recorrente_1_compra": np.random.choice([0, 1], n),
        "Produto Fonte": np.random.choice(["Excel", "Power BI", "Python"], n),
        "Fonte Campanha": np.random.choice(["Google", "Facebook", "Instagram"], n),
        "Sexo": np.random.choice(["Masculino", "Feminino", "Outros"], n),
        "Formacao": np.random.choice(["Superior", "Médio", "Técnico"], n),
        "mes_compra": np.random.randint(1, 13, n),
        "dia_semana_compra": np.random.randint(0, 7, n)
    })
    for col in ["Produto Fonte", "Fonte Campanha", "Sexo", "Formacao"]:
        df[col] = df[col].astype("string")
    df["recorrente_1_compra"] = df["recorrente_1_compra"].astype("Int64")
    return df

def test_predict_linear_formula_python_returns_float(dummy_final_df):
    artifacts = fit_linear_calculator_artifacts(dummy_final_df, random_state=42)
    client = default_test_clients()[0]

    prediction = predict_linear_formula_python(artifacts, client)

    assert isinstance(prediction, float)

def test_create_ltv_calculator_workbook_creates_expected_structure(dummy_final_df, tmp_path: Path):
    artifacts = fit_linear_calculator_artifacts(dummy_final_df, random_state=42)
    output_path = tmp_path / "ltv_calculator.xlsx"

    created_path = create_ltv_calculator_workbook(artifacts, output_path)

    assert created_path.exists()

    workbook = load_workbook(created_path, data_only=False)
    assert workbook.sheetnames == ["Calculadora", "Apoio", "Documentacao"]
    assert "LTV_Previsto" in workbook.defined_names
    assert "CAC_Maximo_Recomendado" in workbook.defined_names
